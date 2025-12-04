import re
import openpyxl
from openpyxl.chart import BarChart, Reference
import math

# Read the text file
with open('file_sizes.txt', 'r') as f:
    lines = f.readlines()

# Extract first non-zero digit from each line
first_digits = []
for line in lines:
    match = re.match(r'(\d+\.?\d*)', line)
    if match:
        number = match.group(1)
        for char in number:
            if char.isdigit() and char != '0':
                first_digits.append(int(char))
                break

total_count = len(first_digits)

# Calculate distribution
distribution = {}
for digit in range(1, 10):
    count = first_digits.count(digit)
    probability = count / total_count if total_count > 0 else 0
    distribution[digit] = {'count': count, 'probability': probability}

# Calculate Benford's Law probabilities
benford_probabilities = {}
for n in range(1, 10):
    benford_probabilities[n] = math.log10(1 + 1/n)

# Create Excel workbook
wb = openpyxl.Workbook()

# Sheet 1: Raw Data
ws1 = wb.active
ws1.title = "Raw First Digits"
ws1['A1'] = "First Non-Zero Digit"
for i, digit in enumerate(first_digits, start=2):
    ws1[f'A{i}'] = digit

# Sheet 2: Distribution & Probability
ws2 = wb.create_sheet("Distribution")
ws2['A1'] = "Digit"
ws2['B1'] = "Count"
ws2['C1'] = "Observed Probability"
for digit in range(1, 10):
    row = digit + 1
    ws2[f'A{row}'] = digit
    ws2[f'B{row}'] = distribution[digit]['count']
    ws2[f'C{row}'] = round(distribution[digit]['probability'], 4)

# Add chart to Distribution sheet
chart2 = BarChart()
chart2.title = "Observed Probability Distribution"
chart2.x_axis.title = "Digit"
chart2.y_axis.title = "Probability"
data2 = Reference(ws2, min_col=3, max_col=3, min_row=1, max_row=10)
cats2 = Reference(ws2, min_col=1, min_row=2, max_row=10)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
ws2.add_chart(chart2, "E2")

# Sheet 3: Benford's Law Calculation
ws3 = wb.create_sheet("Benford's Law")
ws3['A1'] = "Digit (n)"
ws3['B1'] = "P(n) = log₁₀(1 + 1/n)"
for n in range(1, 10):
    row = n + 1
    ws3[f'A{row}'] = n
    ws3[f'B{row}'] = round(benford_probabilities[n], 4)

# Add chart to Benford's Law sheet
chart3 = BarChart()
chart3.title = "Benford's Law Probability Distribution"
chart3.x_axis.title = "Digit"
chart3.y_axis.title = "Probability P(n)"
data3 = Reference(ws3, min_col=2, max_col=2, min_row=1, max_row=10)
cats3 = Reference(ws3, min_col=1, min_row=2, max_row=10)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws3.add_chart(chart3, "D2")

# Sheet 4: Comparison
ws4 = wb.create_sheet("Comparison")
ws4['A1'] = "Digit"
ws4['B1'] = "Observed Probability"
ws4['C1'] = "Benford's Law P(n)"
ws4['D1'] = "Difference"
for digit in range(1, 10):
    row = digit + 1
    ws4[f'A{row}'] = digit
    obs_prob = distribution[digit]['probability']
    benford_prob = benford_probabilities[digit]
    ws4[f'B{row}'] = round(obs_prob, 4)
    ws4[f'C{row}'] = round(benford_prob, 4)
    ws4[f'D{row}'] = round(obs_prob - benford_prob, 4)

# Add chart to comparison sheet
chart = BarChart()
chart.title = "Observed vs Benford's Law"
chart.x_axis.title = "Digit"
chart.y_axis.title = "Probability"

# Data for chart
data = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=10)
cats = Reference(ws4, min_col=1, min_row=2, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws4.add_chart(chart, "F2")

# Save the file
wb.save('Stats_Tables_And_Graph.xlsx')

print(f"Excel file created with {total_count} first digits!")
print(f"\nDigit distribution:")
for digit in range(1, 10):
    count = distribution[digit]['count']
    prob = distribution[digit]['probability']
    benford = benford_probabilities[digit]
    print(f"  Digit {digit}: Count={count:3d}, Observed P={prob:.4f}, Benford P={benford:.4f}")